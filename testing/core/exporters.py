from __future__ import annotations

from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .model import Record, Issue, Severity


def export_references_txt(text: str, path: str | Path) -> None:
    p = Path(path)
    p.write_text(text.rstrip() + "\n", encoding="utf-8")


def _autosize(ws, max_width: int = 70) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            s = str(v)
            max_len = max(max_len, len(s))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len + 2, max_width))


def _style_header(ws, header_row: int = 1) -> None:
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
    header_align = Alignment(vertical="top", wrap_text=True)

    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.freeze_panes = "A2"


def export_records_xlsx(records: List[Record], path: str | Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "records"

    headers = [
        "record_id",
        "type",
        "title",
        "year",
        "authors",
        "container_title",
        "volume",
        "issue",
        "pages",
        "doi",
        "url",
        "publisher",
        "institution",
        "source_file",
        "source_index",
        "dirty",
    ]
    ws.append(headers)

    wrap = Alignment(vertical="top", wrap_text=True)

    for r in records:
        authors = "; ".join(a.display() for a in r.authors) if r.authors else ""
        ws.append([
            r.id,
            r.type.value,
            r.title or "",
            r.year or "",
            authors,
            r.container_title or "",
            r.volume or "",
            r.issue or "",
            r.pages or "",
            r.doi or "",
            r.url or "",
            r.publisher or "",
            r.institution or "",
            r.source_file or "",
            r.source_record_index if r.source_record_index is not None else "",
            "Y" if r.dirty else "",
        ])

    _style_header(ws)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = wrap

    _autosize(ws)
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


def export_issues_xlsx(records: List[Record], global_issues: List[Issue], path: str | Path) -> None:
    wb = Workbook()

    # Sheet 1: record issues
    ws1 = wb.active
    ws1.title = "record_issues"
    ws1.append(["severity", "record_id", "field", "message", "suggestions", "source_file", "title"])

    wrap = Alignment(vertical="top", wrap_text=True)

    # Map record_id -> record (for context columns)
    rec_map = {r.id: r for r in records}

    for r in records:
        for it in r.issues:
            rr = rec_map.get(it.record_id or r.id, r)
            ws1.append([
                it.severity.value,
                it.record_id or r.id,
                it.field,
                it.message,
                "; ".join(it.suggestions) if it.suggestions else "",
                rr.source_file or "",
                rr.title or "",
            ])

    _style_header(ws1)
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
        for cell in row:
            cell.alignment = wrap
    _autosize(ws1)

    # Sheet 2: global issues (file parse errors etc.)
    ws2 = wb.create_sheet("global_issues")
    ws2.append(["severity", "field", "message", "code"])
    for it in global_issues:
        ws2.append([it.severity.value, it.field, it.message, it.code or ""])
    _style_header(ws2)
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=ws2.max_column):
        for cell in row:
            cell.alignment = wrap
    _autosize(ws2)

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
